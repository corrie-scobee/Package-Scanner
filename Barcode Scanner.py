import tkinter as tk
from tkinter import ttk
import openpyxl
from openpyxl import load_workbook
import datetime
from tkinter import font
import tkinter.messagebox
import os
from openpyxl.styles import Alignment


def create_new_excel_for_month():
    current_date = datetime.datetime.now()
    current_month_year = current_date.strftime("%m-%Y")
    new_excel_filename = f"Tracking Numbers {current_month_year}.xlsx"
    if not os.path.exists(new_excel_filename):
        workbook = openpyxl.Workbook()
        workbook.save(new_excel_filename)
    return new_excel_filename

def save_button_click(entries):
    new_excel_filename = create_new_excel_for_month()

    # Try to load the existing workbook
    try:
        workbook = load_workbook(new_excel_filename)
        sheet = workbook.active
    except FileNotFoundError:
        # Create a new workbook if the file doesn't exist
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Set column widths (if needed)
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 5
    sheet.column_dimensions['C'].width = 21

    # Find the next available row
    # next_row = sheet.max_row + 1
    next_row = 1
    while sheet.cell(row=next_row, column=1).value is not None:
        next_row += 1

    temp_date = datetime.datetime.now().date()
    current_date = temp_date.strftime("%m-%d-%Y")

    # Write the data from entry boxes to the next available row
    for entry in entries:
        if isinstance(entry, tk.Entry):
            entry_value = entry.get()
            if entry_value:
                sheet.cell(row=next_row, column=3, value=entry_value)
                sheet.cell(row=next_row, column=1, value=current_date)
                next_row += 1

    for row in sheet.iter_rows(min_row=1, max_row=next_row, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save the Excel file
    workbook.save(new_excel_filename)

    tkinter.messagebox.showinfo("Notification", "Entries Submitted! \n\nYou can now exit the app or add new entries.")

    for entry in entries:
        if isinstance(entry, tk.Entry):
            entry.delete(0, tk.END)

    if entries and isinstance(entries[0], tk.Entry):
        entries[0].focus_set()  


def validate_barcode(barcode):
# Assuming UPS tracking number has 18 characters
    ups_length = 18
    # Assuming FedEx tracking number has 34 characters
    fedex_length = 34
    new_barcode = str(barcode)
    if len(new_barcode) == ups_length:
    # UPS tracking number, no adjustment needed
        return barcode
    elif len(new_barcode) == fedex_length and new_barcode.isdigit():
        # FedEx tracking number with 34 characters
        # Remove the first 22 digits as FedEx tracking number
        return new_barcode[22:]
    else:
        # Invalid barcode
        return barcode

def focus_next_entry(event):
    current_entry = event.widget
    current_text = current_entry.get()

    # Check if the last character is a delimiter (e.g., Enter or Tab)
    if event.keysym == 'Return':
        # Validate and adjust the barcode
        validated_barcode = validate_barcode(current_text)
        current_entry.delete(0, tk.END)
        current_entry.insert(0, validated_barcode)
        current_entry.tk_focusNext().focus()

    # Collect all validated barcodes
    validated_barcodes = [validate_barcode(entry.get()) for entry in entries]

    # Remove None values from the list
    validated_barcodes = [barcode for barcode in validated_barcodes if barcode !='']

    # save_button_click(validated_barcodes)

# Create the main window
app = tk.Tk()
app.title("Package Tracker")
app.geometry("500x600")

label1_font = font.Font(family="Helvetica", size=25)

# Create and add widgets (e.g., labels, buttons, entry fields)
label = tk.Label(app, text="Package Tracker", font = label1_font)
label.place(relx=0.5, rely=0.05, anchor=tk.CENTER)

num_entries = 40
entries = [tk.Entry(app) for _ in range(num_entries)]

# Place the initial entries using place
for i, entry in enumerate(entries):
    entry.place(relx=0.3 + i % 2 * 0.4, rely=i // 2 * 0.04 + 0.2, anchor=tk.CENTER)
    entry.bind("<Key>", focus_next_entry)

entries[0].focus_set()


save_button = ttk.Button(app, text="Save", command=lambda: save_button_click(entries))
save_button.place(relx=0.5, rely=0.12, anchor=tk.CENTER)

# sv_ttk.set_theme("light")

# Run the main loop
app.mainloop()